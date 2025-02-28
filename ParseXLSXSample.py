import json
import os
import yaml
import math
import numpy as np
from math import cos,sin,pi,sqrt
from numpy import deg2rad, rad2deg
import openpyxl
from chemformula import ChemFormula
import inspect

def ExtractRunNumber(fname):
	splitted=os.path.basename(fname).split('_')
	if(len(splitted)>0):
		str_number=splitted[0]
		str_number=str_number.replace('test','')
		return int(str_number)
	return -1

def UpdateLocalSampleTable(Username,IP,Port):
	LocalPath=os.path.dirname(inspect.getfile(inspect.currentframe()))
	os.system("scp -P %d %s@%s:/RAID1/SAMPLE_information/samples_standard_boxes_info.xlsx %s" % (Port,Username,IP,LocalPath))
	os.system("scp -P %d %s@%s:/RAID1/SAMPLE_information/sample_data.xlsx %s" % (Port,Username,IP,LocalPath))

def TestNumbersFromTable(value):
	result=[]
	if not value:
		return []
	if str(value).find('-') > 0:
		splitted=value.split('-')
		#print(splitted)
		try:
			for i in range(int(splitted[0]),int(splitted[1])+1):
				result.append(int(i))
			return result
		except:
			return []
	else:
		splitted=str(value).split(',')
		#print(splitted)
		for i in splitted:
			try:
				result.append(int(i))
			except:
				continue
		return result
	return []


def ReadSampleDataFile(filename=""):
	if len(filename)==0:
		LocalPath=os.path.dirname(inspect.getfile(inspect.currentframe()))+"/sample_data.xlsx"
		#print(LocalPath)
		if os.path.isfile(LocalPath):
			filename=LocalPath
	workbook = openpyxl.load_workbook(filename,data_only=True)
	worksheet = workbook['Sheet']
	dict_res={}
	#берем ключи из файла:
	keys=[]
	for col in range(1,worksheet.max_column + 1):
		keys.append(str(worksheet.cell(1, col).value))
	for row in range(2, worksheet.max_row + 1):
		dict_row={}
		for col in  range(1,worksheet.max_column + 1):
			value=str(worksheet.cell(row, col).value)
			if value.find("[") > -1:
				value=value.replace("[","")
				value=value.replace("]","")
				splitted=value.split(",")
				dict_row[keys[col-1]]=[]
				for it in splitted:
					try:
						dict_row[keys[col-1]].append(float(it))
					except:
						dict_row[keys[col-1]].append(it)
			else:
				try:
					dict_row[keys[col-1]]=float(value)
				except:
					dict_row[keys[col-1]]=value
					if dict_row[keys[col-1]]=="True":
						dict_row[keys[col-1]]=True
					elif dict_row[keys[col-1]]=="False":
						dict_row[keys[col-1]]=False
		dict_res[dict_row["Filename"]]=dict(dict_row)
	return dict_res

def ReadSampleFile(filename=""):
	AverageBoxMass={10: 341, 20: 365, 30: 408, 40: 466}
	if len(filename)==0:
		LocalPath=os.path.dirname(inspect.getfile(inspect.currentframe()))+"/samples_standard_boxes_info.xlsx"
		#print(LocalPath)
		if os.path.isfile(LocalPath):
			filename=LocalPath
	workbook = openpyxl.load_workbook(filename,data_only=True)
	worksheet = workbook['SampleBox']
	dict_res={}
	for row in range(2, worksheet.max_row + 1):
		dict_row={}
		key = str(worksheet.cell(row, 1).value)
		if key:
			if key.isnumeric():
				#print(key)
				try:
					#print(worksheet.cell(row, 11).value)
					Bad=worksheet.cell(row, 11).value
					BadRuns=TestNumbersFromTable(Bad)
					
					dict_row['Sample']=worksheet.cell(row, 2).value
					dict_row['Thickness']=worksheet.cell(row, 3).value
					dict_row['Mass']=float(worksheet.cell(row, 5).value)
					dict_row['Density']=worksheet.cell(row, 6).value
					dict_row['MolarMass']=0
					dict_row['SourceCoordinates']=[0,0,-7.5] # координаты источника относительно центра системы
					dict_row['SampleCoordinates']=[0,0,-7.5]
					PosZ_str=str(worksheet.cell(row, 8).value)
					PosY_str=str(worksheet.cell(row, 9).value)
					if PosZ_str.find('34')>-1:#образец стоял далеко от генератора
						dict_row['SourceCoordinates'][0]=391.5# нашел ошибку при перепроверке
						dict_row['SampleCoordinates'][0]=386-dict_row['Thickness']/2
						dict_row['Displaced']=True
					else:
						dict_row['SourceCoordinates'][0]=345.5+dict_row['Thickness']+1.5
						dict_row['SampleCoordinates'][0]=338+dict_row['Thickness']/2
						dict_row['Displaced']=False
					PosY_str=PosY_str.replace(' move box','')
					dict_row['SourceCoordinates'][2]=float(PosY_str)
					BoxMass=worksheet.cell(row, 4).value
					#print(BoxMass)
					if str(BoxMass).isnumeric():
						dict_row['BoxMass']=float(BoxMass)
					else:
						dict_row['BoxMass']=float(AverageBoxMass[int(dict_row['Thickness'])])
					#dict_row['PositionZ']=worksheet.cell(row, 3).value
					runs=TestNumbersFromTable(str(worksheet.cell(row, 7).value))
					#print(dict_row['Sample'])
					#print(dict_row)
					if not (dict_row['Sample']=='Bskgr'):
						formula=ChemFormula(dict_row['Sample'])
						dict_row['MolarMass']=formula.formula_weight
						dict_row['Elements']=[]
						dict_row['NAtoms']=[]
						for i in formula.element.keys():
							dict_row['Elements'].append(i)
							dict_row['NAtoms'].append(formula.element[i])
						#print(formula.element,formula.formula_weight)
					else:
						dict_row['MolarMass']=0
						dict_row['Elements']=[0]
						dict_row['NAtoms']=[0]
						dict_row['Elements'].append("")
						dict_row['NAtoms'].append(0)
					#print(dict_row)
					for i in runs:
						if i in BadRuns:
							dict_row['Bad']=True
						else:
							dict_row['Bad']=False
						dict_res[i]=dict(dict_row)
							#print(dict_row)
							
				except:
					continue
	return dict_res
